import streamlit as st
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from validation_editor import ValidationEditor

class ErrorReporter:
    def __init__(self):
        self.validation_editor = ValidationEditor()
        self.error_categories = {
            'ocr_quality': {
                'name': 'Calidad OCR',
                'description': 'Problemas relacionados con la extracción de texto',
                'color': 'FF6B6B',
                'icon': ''
            },
            'field_validation': {
                'name': 'Validación de Campos',
                'description': 'Errores en la validación de datos extraídos',
                'color': 'FFE66D',
                'icon': ''
            },
            'document_structure': {
                'name': 'Estructura del Documento',
                'description': 'Problemas con el formato o tipo de documento',
                'color': 'FF8E53',
                'icon': ''
            },
            'data_completeness': {
                'name': 'Completitud de Datos',
                'description': 'Campos faltantes o incompletos',
                'color': '95E1D3',
                'icon': ''
            },
            'calculation_errors': {
                'name': 'Errores de Cálculo',
                'description': 'Problemas en cálculos automáticos',
                'color': 'B19CD9',
                'icon': ''
            }
        }
        
        self.severity_levels = {
            'critical': {'name': 'Crítico', 'color': 'FF0000', 'weight': 4},
            'high': {'name': 'Alto', 'color': 'FF8000', 'weight': 3},
            'medium': {'name': 'Medio', 'color': 'FFD700', 'weight': 2},
            'low': {'name': 'Bajo', 'color': '90EE90', 'weight': 1}
        }
    
    def analyze_processing_errors(self, processed_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze all types of errors in processed documents"""
        analysis = {
            'total_documents': len(processed_data),
            'documents_with_errors': 0,
            'error_summary': {category: 0 for category in self.error_categories.keys()},
            'severity_summary': {level: 0 for level in self.severity_levels.keys()},
            'detailed_errors': [],
            'quality_metrics': {},
            'recommendations': [],
            'processing_statistics': {}
        }
        
        for doc_index, document in enumerate(processed_data):
            doc_errors = self._analyze_document_errors(document, doc_index)
            
            if doc_errors['errors']:
                analysis['documents_with_errors'] += 1
                analysis['detailed_errors'].extend(doc_errors['errors'])
                
                # Update category and severity counts
                for error in doc_errors['errors']:
                    analysis['error_summary'][error['category']] += 1
                    analysis['severity_summary'][error['severity']] += 1
        
        # Calculate quality metrics
        analysis['quality_metrics'] = self._calculate_quality_metrics(processed_data, analysis)
        
        # Generate recommendations
        analysis['recommendations'] = self._generate_recommendations(analysis, processed_data)
        
        # Processing statistics
        analysis['processing_statistics'] = self._get_processing_statistics(processed_data)
        
        return analysis
    
    def _analyze_document_errors(self, document: Dict[str, Any], doc_index: int) -> Dict[str, Any]:
        """Analyze errors in a single document"""
        doc_analysis = {
            'document_index': doc_index,
            'filename': document.get('archivo', f'Documento_{doc_index + 1}'),
            'errors': [],
            'warnings': []
        }
        
        # OCR Quality Analysis
        ocr_errors = self._analyze_ocr_quality(document)
        doc_analysis['errors'].extend(ocr_errors)
        
        # Field Validation Analysis
        validation_errors = self._analyze_field_validation(document)
        doc_analysis['errors'].extend(validation_errors)
        
        # Document Structure Analysis
        structure_errors = self._analyze_document_structure(document)
        doc_analysis['errors'].extend(structure_errors)
        
        # Data Completeness Analysis
        completeness_errors = self._analyze_data_completeness(document)
        doc_analysis['errors'].extend(completeness_errors)
        
        # Calculation Errors Analysis
        calculation_errors = self._analyze_calculation_errors(document)
        doc_analysis['errors'].extend(calculation_errors)
        
        return doc_analysis
    
    def _analyze_ocr_quality(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze OCR quality issues"""
        errors = []
        
        # Check processing method
        processing_method = document.get('metodo_procesamiento', '')
        
        # Check for fallback processing (indicates OCR issues)
        if processing_method == 'fallback':
            errors.append({
                'category': 'ocr_quality',
                'severity': 'critical',
                'message': 'Documento requiere conversión a imagen para procesamiento OCR',
                'field': 'metodo_procesamiento',
                'suggested_action': 'Convertir PDF a imagen de alta resolución antes del procesamiento',
                'impact': 'No se pudieron extraer datos del documento'
            })
        
        # Check for OCR confidence indicators in errors
        doc_errors = document.get('errores', [])
        if isinstance(doc_errors, list):
            for error_msg in doc_errors:
                if any(keyword in str(error_msg).lower() for keyword in ['texto', 'ilegible', 'borroso', 'calidad']):
                    errors.append({
                        'category': 'ocr_quality',
                        'severity': 'high',
                        'message': f'Problema de calidad en OCR: {error_msg}',
                        'field': 'general',
                        'suggested_action': 'Mejorar calidad de imagen o usar documento original',
                        'impact': 'Datos extraídos pueden ser incorrectos'
                    })
        
        # Analyze field extraction quality
        required_fields = ['numero_bl', 'exportador', 'consignatario']
        missing_critical = []
        
        for field in required_fields:
            value = document.get(field, 'No detectado')
            if value == 'No detectado' or not value:
                missing_critical.append(field)
        
        if len(missing_critical) >= 2:  # Multiple critical fields missing
            errors.append({
                'category': 'ocr_quality',
                'severity': 'high',
                'message': f'Múltiples campos críticos no detectados: {", ".join(missing_critical)}',
                'field': 'multiple',
                'suggested_action': 'Verificar calidad de imagen y orientación del documento',
                'impact': 'Documento puede no ser procesable correctamente'
            })
        
        return errors
    
    def _analyze_field_validation(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze field validation errors"""
        errors = []
        
        # Use ValidationEditor to check field validation
        validation_result = self.validation_editor.validate_document(document)
        
        if not validation_result['valid']:
            for field_name, field_result in validation_result['fields'].items():
                if not field_result['valid']:
                    field_display_name = self.validation_editor.field_definitions[field_name]['name']
                    severity = 'critical' if self.validation_editor.field_definitions[field_name].get('required', False) else 'medium'
                    
                    for message in field_result['messages']:
                        errors.append({
                            'category': 'field_validation',
                            'severity': severity,
                            'message': f'{field_display_name}: {message}',
                            'field': field_name,
                            'suggested_action': 'Revisar y corregir el valor manualmente',
                            'impact': 'Datos incorrectos en el reporte final' if severity == 'medium' else 'Campo obligatorio faltante'
                        })
        
        return errors
    
    def _analyze_document_structure(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze document structure issues"""
        errors = []
        
        # Check document type detection
        doc_type = document.get('tipo_documento', '')
        if not doc_type or doc_type == 'PDF_SIN_TEXTO':
            errors.append({
                'category': 'document_structure',
                'severity': 'high',
                'message': 'Tipo de documento no identificado o sin texto extraíble',
                'field': 'tipo_documento',
                'suggested_action': 'Verificar que el documento sea un BL o factura válida',
                'impact': 'Procesamiento incompleto del documento'
            })
        
        # Check for mixed document indicators
        has_bl_fields = any(document.get(field) not in ['No detectado', None, ''] 
                           for field in ['numero_bl', 'numero_contenedor'])
        has_invoice_fields = any(document.get(field) not in ['No detectado', None, ''] 
                               for field in ['valor_factura', 'incoterm'])
        
        if not has_bl_fields and not has_invoice_fields:
            errors.append({
                'category': 'document_structure',
                'severity': 'critical',
                'message': 'No se detectaron campos típicos de BL ni de factura',
                'field': 'general',
                'suggested_action': 'Verificar que el documento sea del tipo correcto',
                'impact': 'Documento no puede ser procesado correctamente'
            })
        
        # Check INCOTERM consistency
        incoterm = document.get('incoterm', '')
        valor_fob = document.get('valor_fob', 'No detectado')
        valor_cif = document.get('valor_cif', 'No detectado')
        
        if incoterm == 'FOB' and valor_fob == 'No detectado':
            errors.append({
                'category': 'document_structure',
                'severity': 'medium',
                'message': 'INCOTERM es FOB pero no se detectó valor FOB',
                'field': 'valor_fob',
                'suggested_action': 'Verificar si el valor FOB está presente en el documento',
                'impact': 'Cálculo CIF puede ser incorrecto'
            })
        
        return errors
    
    def _analyze_data_completeness(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze data completeness issues"""
        errors = []
        
        # Define field importance levels
        critical_fields = ['numero_bl', 'exportador', 'consignatario']
        important_fields = ['valor_flete', 'bultos', 'kilos', 'incoterm']
        optional_fields = ['numero_contenedor', 'valor_factura', 'moneda']
        
        # Check critical fields
        missing_critical = []
        for field in critical_fields:
            value = document.get(field, 'No detectado')
            if value in ['No detectado', None, '']:
                missing_critical.append(self.validation_editor.field_definitions[field]['name'])
        
        if missing_critical:
            errors.append({
                'category': 'data_completeness',
                'severity': 'critical',
                'message': f'Campos críticos faltantes: {", ".join(missing_critical)}',
                'field': 'multiple',
                'suggested_action': 'Completar manualmente los campos críticos',
                'impact': 'El documento no puede ser procesado sin estos campos'
            })
        
        # Check important fields
        missing_important = []
        for field in important_fields:
            value = document.get(field, 'No detectado')
            if value in ['No detectado', None, '']:
                missing_important.append(self.validation_editor.field_definitions[field]['name'])
        
        if len(missing_important) > len(important_fields) / 2:  # More than half missing
            errors.append({
                'category': 'data_completeness',
                'severity': 'medium',
                'message': f'Múltiples campos importantes faltantes: {", ".join(missing_important)}',
                'field': 'multiple',
                'suggested_action': 'Revisar el documento para completar campos faltantes',
                'impact': 'Información incompleta para análisis y reportes'
            })
        
        # Calculate completeness percentage
        all_fields = critical_fields + important_fields + optional_fields
        detected_fields = sum(1 for field in all_fields 
                            if document.get(field, 'No detectado') not in ['No detectado', None, ''])
        completeness = (detected_fields / len(all_fields)) * 100
        
        if completeness < 50:
            errors.append({
                'category': 'data_completeness',
                'severity': 'high',
                'message': f'Completitud de datos muy baja: {completeness:.1f}%',
                'field': 'general',
                'suggested_action': 'Revisar calidad del documento y procesar nuevamente',
                'impact': 'Datos insuficientes para generar reporte confiable'
            })
        
        return errors
    
    def _analyze_calculation_errors(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze calculation and conversion errors"""
        errors = []
        
        # Check FOB to CIF conversion
        conversion_attempted = document.get('conversion_fob_cif', False)
        if conversion_attempted:
            valor_fob = document.get('valor_fob', 'No detectado')
            valor_cif = document.get('valor_cif', 'No detectado')
            
            try:
                fob_num = float(str(valor_fob).replace(',', '').replace('$', ''))
                cif_num = float(str(valor_cif).replace(',', '').replace('$', ''))
                
                # Check if conversion ratio is reasonable (CIF should be higher than FOB)
                if cif_num <= fob_num:
                    errors.append({
                        'category': 'calculation_errors',
                        'severity': 'medium',
                        'message': 'Valor CIF no es mayor que FOB después de conversión',
                        'field': 'valor_cif',
                        'suggested_action': 'Verificar valores FOB y CIF manualmente',
                        'impact': 'Cálculo de conversión puede ser incorrecto'
                    })
                
                # Check if conversion ratio is too high (suspicious)
                ratio = cif_num / fob_num if fob_num > 0 else 0
                if ratio > 1.5:  # More than 50% increase
                    errors.append({
                        'category': 'calculation_errors',
                        'severity': 'low',
                        'message': f'Ratio CIF/FOB muy alto: {ratio:.2f} (esperado ~1.15)',
                        'field': 'conversion_fob_cif',
                        'suggested_action': 'Verificar si los costos de flete y seguro son correctos',
                        'impact': 'Conversión FOB a CIF puede estar sobrestimada'
                    })
            
            except (ValueError, TypeError):
                errors.append({
                    'category': 'calculation_errors',
                    'severity': 'medium',
                    'message': 'No se pudo validar conversión FOB a CIF por valores no numéricos',
                    'field': 'conversion_fob_cif',
                    'suggested_action': 'Verificar que los valores FOB y CIF sean números válidos',
                    'impact': 'Conversión automática falló'
                })
        
        # Check for inconsistent numeric values
        numeric_fields = ['valor_flete', 'bultos', 'kilos', 'valor_factura', 'valor_fob', 'valor_cif']
        for field in numeric_fields:
            value = document.get(field, 'No detectado')
            if value not in ['No detectado', None, '']:
                try:
                    num_value = float(str(value).replace(',', '').replace('$', ''))
                    if num_value < 0:
                        errors.append({
                            'category': 'calculation_errors',
                            'severity': 'medium',
                            'message': f'{self.validation_editor.field_definitions[field]["name"]} tiene valor negativo',
                            'field': field,
                            'suggested_action': 'Corregir el valor manualmente',
                            'impact': 'Valor inválido en cálculos'
                        })
                except (ValueError, TypeError):
                    errors.append({
                        'category': 'calculation_errors',
                        'severity': 'low',
                        'message': f'{self.validation_editor.field_definitions[field]["name"]} no es un número válido',
                        'field': field,
                        'suggested_action': 'Convertir a formato numérico correcto',
                        'impact': 'Campo no puede ser usado en cálculos'
                    })
        
        return errors
    
    def _calculate_quality_metrics(self, processed_data: List[Dict[str, Any]], analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate overall quality metrics"""
        total_docs = len(processed_data)
        if total_docs == 0:
            return {}
        
        # Calculate error rates
        error_rate = (analysis['documents_with_errors'] / total_docs) * 100
        success_rate = 100 - error_rate
        
        # Calculate field detection rates
        all_fields = list(self.validation_editor.field_definitions.keys())
        field_detection_rates = {}
        
        for field in all_fields:
            detected_count = sum(1 for doc in processed_data 
                               if doc.get(field, 'No detectado') not in ['No detectado', None, ''])
            field_detection_rates[field] = (detected_count / total_docs) * 100
        
        # Calculate severity distribution
        total_errors = sum(analysis['severity_summary'].values())
        severity_distribution = {}
        for severity, count in analysis['severity_summary'].items():
            severity_distribution[severity] = (count / total_errors * 100) if total_errors > 0 else 0
        
        # Calculate average completeness
        completeness_scores = []
        for doc in processed_data:
            detected_fields = sum(1 for field in all_fields 
                                if doc.get(field, 'No detectado') not in ['No detectado', None, ''])
            completeness = (detected_fields / len(all_fields)) * 100
            completeness_scores.append(completeness)
        
        avg_completeness = sum(completeness_scores) / len(completeness_scores) if completeness_scores else 0
        
        return {
            'success_rate': round(success_rate, 2),
            'error_rate': round(error_rate, 2),
            'average_completeness': round(avg_completeness, 2),
            'field_detection_rates': {field: round(rate, 1) for field, rate in field_detection_rates.items()},
            'severity_distribution': {sev: round(dist, 1) for sev, dist in severity_distribution.items()},
            'quality_score': round((success_rate + avg_completeness) / 2, 2)  # Combined quality score
        }
    
    def _generate_recommendations(self, analysis: Dict[str, Any], processed_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Generate actionable recommendations based on error analysis"""
        recommendations = []
        
        # High error rate recommendations
        if analysis['quality_metrics'].get('error_rate', 0) > 50:
            recommendations.append({
                'priority': 'critical',
                'category': 'general',
                'title': 'Alta tasa de errores detectada',
                'description': 'Más del 50% de los documentos tienen errores',
                'actions': [
                    'Verificar calidad de las imágenes/PDFs',
                    'Asegurar que los documentos estén bien orientados',
                    'Revisar que sean documentos BL o facturas válidas'
                ]
            })
        
        # OCR quality recommendations
        ocr_errors = analysis['error_summary'].get('ocr_quality', 0)
        if ocr_errors > 0:
            recommendations.append({
                'priority': 'high',
                'category': 'ocr_quality',
                'title': 'Problemas de calidad OCR detectados',
                'description': f'{ocr_errors} errores relacionados con extracción de texto',
                'actions': [
                    'Usar imágenes de mayor resolución (mínimo 300 DPI)',
                    'Asegurar buen contraste entre texto y fondo',
                    'Convertir PDFs escaneados a imágenes de alta calidad'
                ]
            })
        
        # Data completeness recommendations
        avg_completeness = analysis['quality_metrics'].get('average_completeness', 0)
        if avg_completeness < 70:
            recommendations.append({
                'priority': 'medium',
                'category': 'data_completeness',
                'title': 'Completitud de datos baja',
                'description': f'Completitud promedio: {avg_completeness:.1f}%',
                'actions': [
                    'Revisar manualmente los documentos con más campos faltantes',
                    'Verificar que todos los campos estén visibles en el documento',
                    'Usar la función de edición manual para completar datos'
                ]
            })
        
        # Field-specific recommendations
        detection_rates = analysis['quality_metrics'].get('field_detection_rates', {})
        low_detection_fields = [field for field, rate in detection_rates.items() if rate < 50]
        
        if low_detection_fields:
            recommendations.append({
                'priority': 'medium',
                'category': 'field_validation',
                'title': 'Campos con baja detección',
                'description': f'Campos problemáticos: {", ".join([self.validation_editor.field_definitions[f]["name"] for f in low_detection_fields])}',
                'actions': [
                    'Verificar ubicación de estos campos en los documentos',
                    'Asegurar que el texto esté claramente legible',
                    'Considerar usar plantillas de documentos estandarizadas'
                ]
            })
        
        return recommendations
    
    def _get_processing_statistics(self, processed_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Get processing method statistics"""
        stats = {
            'processing_methods': {},
            'document_types': {},
            'incoterms': {},
            'currencies': {}
        }
        
        for doc in processed_data:
            # Processing methods
            method = doc.get('metodo_procesamiento', 'unknown')
            stats['processing_methods'][method] = stats['processing_methods'].get(method, 0) + 1
            
            # Document types
            doc_type = doc.get('tipo_documento', 'unknown')
            stats['document_types'][doc_type] = stats['document_types'].get(doc_type, 0) + 1
            
            # INCOTERMs
            incoterm = doc.get('incoterm', 'No detectado')
            if incoterm != 'No detectado':
                stats['incoterms'][incoterm] = stats['incoterms'].get(incoterm, 0) + 1
            
            # Currencies
            currency = doc.get('moneda', 'No detectado')
            if currency != 'No detectado':
                stats['currencies'][currency] = stats['currencies'].get(currency, 0) + 1
        
        return stats
    
    def render_error_dashboard(self, processed_data: List[Dict[str, Any]]):
        """Render comprehensive error analysis dashboard"""
        st.header("Reporte Detallado de Errores y Calidad")
        
        if not processed_data:
            st.warning("No hay datos procesados para analizar")
            return
        
        # Perform analysis
        with st.spinner("Analizando errores y calidad de datos..."):
            analysis = self.analyze_processing_errors(processed_data)
        
        # Overview metrics
        self._render_overview_metrics(analysis)
        
        # Detailed sections in tabs
        tab1, tab2, tab3, tab4 = st.tabs([
            "Errores Detallados", 
            "Métricas de Calidad", 
            "Recomendaciones",
            "Estadísticas"
        ])
        
        with tab1:
            self._render_detailed_errors(analysis)
        
        with tab2:
            self._render_quality_metrics(analysis)
        
        with tab3:
            self._render_recommendations(analysis)
        
        with tab4:
            self._render_processing_statistics(analysis)
    
    def _render_overview_metrics(self, analysis: Dict[str, Any]):
        """Render overview metrics at the top"""
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            success_rate = analysis['quality_metrics'].get('success_rate', 0)
            st.metric(
                label="Tasa de Éxito",
                value=f"{success_rate}%",
                delta=f"{100 - success_rate:.1f}% errores" if success_rate < 100 else "Sin errores"
            )
        
        with col2:
            total_errors = sum(analysis['error_summary'].values())
            st.metric(
                label="Total de Errores",
                value=total_errors,
                delta=f"{analysis['documents_with_errors']} documentos afectados"
            )
        
        with col3:
            avg_completeness = analysis['quality_metrics'].get('average_completeness', 0)
            st.metric(
                label="Completitud Promedio",
                value=f"{avg_completeness:.1f}%",
                delta="de campos detectados"
            )
        
        with col4:
            quality_score = analysis['quality_metrics'].get('quality_score', 0)
            color = "normal" if quality_score >= 80 else "inverse"
            st.metric(
                label="Puntuación de Calidad",
                value=f"{quality_score:.1f}",
                delta="sobre 100",
                delta_color=color
            )
    
    def _render_detailed_errors(self, analysis: Dict[str, Any]):
        """Render detailed error breakdown"""
        st.subheader("Análisis Detallado de Errores")
        
        if not analysis['detailed_errors']:
            st.success("No se encontraron errores en los documentos procesados")
            return
        
        # Error summary by category
        st.markdown("### Resumen por Categoría")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Category breakdown
            for category, count in analysis['error_summary'].items():
                if count > 0:
                    cat_info = self.error_categories[category]
                    st.markdown(f"""
                    <div style="padding: 10px; border-left: 4px solid #{cat_info['color']}; background-color: #f9f9f9; margin: 5px 0;">
                        <strong>{cat_info['name']}</strong><br>
                        <small>{cat_info['description']}</small><br>
                        <span style="color: #{cat_info['color']}; font-weight: bold;">{count} errores</span>
                    </div>
                    """, unsafe_allow_html=True)
        
        with col2:
            # Severity breakdown
            st.markdown("**Distribución por Severidad:**")
            for severity, count in analysis['severity_summary'].items():
                if count > 0:
                    sev_info = self.severity_levels[severity]
                    st.markdown(f"""
                    <div style="padding: 5px; border-radius: 5px; background-color: #{sev_info['color']}20; margin: 2px 0;">
                        <span style="color: #{sev_info['color']}; font-weight: bold;">
                            {sev_info['name']}: {count} errores
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
        
        # Detailed error list
        st.markdown("### Lista Detallada de Errores")
        
        # Group errors by document
        errors_by_doc = {}
        for error in analysis['detailed_errors']:
            doc_key = f"Documento {error.get('document_index', 0) + 1}"
            if doc_key not in errors_by_doc:
                errors_by_doc[doc_key] = []
            errors_by_doc[doc_key].append(error)
        
        for doc_name, doc_errors in errors_by_doc.items():
            with st.expander(f"{doc_name} ({len(doc_errors)} errores)", expanded=False):
                for i, error in enumerate(doc_errors, 1):
                    cat_info = self.error_categories[error['category']]
                    sev_info = self.severity_levels[error['severity']]
                    
                    st.markdown(f"""
                    **{i}. {error['message']}**
                    
                    - **Categoría:** {cat_info['name']}
                    - **Severidad:** <span style="color: #{sev_info['color']}; font-weight: bold;">{sev_info['name']}</span>
                    - **Campo afectado:** {error['field']}
                    - **Acción sugerida:** {error['suggested_action']}
                    - **Impacto:** {error['impact']}
                    """, unsafe_allow_html=True)
                    st.markdown("---")
    
    def _render_quality_metrics(self, analysis: Dict[str, Any]):
        """Render quality metrics visualization"""
        st.subheader("Métricas de Calidad Detalladas")
        
        metrics = analysis['quality_metrics']
        
        # Field detection rates
        st.markdown("### Tasas de Detección por Campo")
        
        detection_data = []
        for field, rate in metrics.get('field_detection_rates', {}).items():
            field_name = self.validation_editor.field_definitions[field]['name']
            detection_data.append({
                'Campo': field_name,
                'Tasa de Detección (%)': rate,
                'Estado': 'Excelente' if rate >= 90 else 'Bueno' if rate >= 70 else 'Necesita Mejora'
            })
        
        if detection_data:
            df_detection = pd.DataFrame(detection_data)
            st.dataframe(df_detection, use_container_width=True)
        
        # Quality distribution
        st.markdown("### Distribución de Calidad")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Métricas Generales:**")
            st.write(f"- **Tasa de éxito:** {metrics.get('success_rate', 0):.1f}%")
            st.write(f"- **Tasa de errores:** {metrics.get('error_rate', 0):.1f}%")
            st.write(f"- **Completitud promedio:** {metrics.get('average_completeness', 0):.1f}%")
            st.write(f"- **Puntuación de calidad:** {metrics.get('quality_score', 0):.1f}/100")
        
        with col2:
            st.markdown("**Distribución de Severidad:**")
            severity_dist = metrics.get('severity_distribution', {})
            for severity, percentage in severity_dist.items():
                if percentage > 0:
                    sev_info = self.severity_levels[severity]
                    st.write(f"- **{sev_info['name']}:** {percentage:.1f}%")
    
    def _render_recommendations(self, analysis: Dict[str, Any]):
        """Render actionable recommendations"""
        st.subheader("Recomendaciones para Mejorar la Calidad")
        
        recommendations = analysis['recommendations']
        
        if not recommendations:
            st.success("No hay recomendaciones específicas. ¡La calidad es excelente!")
            return
        
        # Sort by priority
        priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        sorted_recommendations = sorted(recommendations, 
                                      key=lambda x: priority_order.get(x['priority'], 4))
        
        for i, rec in enumerate(sorted_recommendations, 1):
            priority_colors = {
                'critical': 'FF0000',
                'high': 'FF8000', 
                'medium': 'FFD700',
                'low': '90EE90'
            }
            
            color = priority_colors.get(rec['priority'], '808080')
            
            st.markdown(f"""
            <div style="padding: 15px; border-left: 5px solid #{color}; background-color: #f9f9f9; margin: 10px 0; border-radius: 5px;">
                <h4 style="margin: 0; color: #{color};">
                    {i}. {rec['title']} 
                    <small style="background-color: #{color}20; padding: 2px 8px; border-radius: 10px; color: #{color}; font-weight: bold;">
                        {rec['priority'].upper()}
                    </small>
                </h4>
                <p style="margin: 10px 0; color: #666;">{rec['description']}</p>
                <strong>Acciones recomendadas:</strong>
                <ul style="margin: 5px 0;">
            """, unsafe_allow_html=True)
            
            for action in rec['actions']:
                st.markdown(f"<li>{action}</li>", unsafe_allow_html=True)
            
            st.markdown("</ul></div>", unsafe_allow_html=True)
    
    def _render_processing_statistics(self, analysis: Dict[str, Any]):
        """Render processing statistics"""
        st.subheader("Estadísticas de Procesamiento")
        
        stats = analysis['processing_statistics']
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Métodos de Procesamiento")
            for method, count in stats['processing_methods'].items():
                method_name = {
                    'texto_extraido': 'Texto Extraído (PDF)',
                    'vision_ocr': 'OCR con Visión (Imagen)',
                    'fallback': 'Procesamiento de Respaldo',
                    'unknown': 'Desconocido'
                }.get(method, method)
                st.write(f"- **{method_name}:** {count} documentos")
            
            st.markdown("### Tipos de Documento")
            for doc_type, count in stats['document_types'].items():
                type_name = {
                    'BL': 'Bill of Lading',
                    'FACTURA': 'Factura',
                    'MIXTO': 'Documento Mixto',
                    'PDF_SIN_TEXTO': 'PDF sin Texto',
                    'unknown': 'Desconocido'
                }.get(doc_type, doc_type)
                st.write(f"- **{type_name}:** {count} documentos")
        
        with col2:
            st.markdown("### INCOTERMs Detectados")
            if stats['incoterms']:
                for incoterm, count in stats['incoterms'].items():
                    st.write(f"- **{incoterm}:** {count} documentos")
            else:
                st.write("No se detectaron INCOTERMs válidos")
            
            st.markdown("### Monedas Detectadas")
            if stats['currencies']:
                for currency, count in stats['currencies'].items():
                    st.write(f"- **{currency}:** {count} documentos")
            else:
                st.write("No se detectaron monedas válidas")
    
    def generate_error_report_excel(self, analysis: Dict[str, Any]) -> io.BytesIO:
        """Generate detailed error report in Excel format"""
        wb = Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # Overview sheet
        ws_overview = wb.create_sheet("Resumen Ejecutivo")
        self._create_overview_sheet(ws_overview, analysis)
        
        # Detailed errors sheet
        ws_errors = wb.create_sheet("Errores Detallados")
        self._create_errors_sheet(ws_errors, analysis)
        
        # Quality metrics sheet
        ws_metrics = wb.create_sheet("Métricas de Calidad")
        self._create_metrics_sheet(ws_metrics, analysis)
        
        # Recommendations sheet
        ws_recommendations = wb.create_sheet("Recomendaciones")
        self._create_recommendations_sheet(ws_recommendations, analysis)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_overview_sheet(self, ws, analysis: Dict[str, Any]):
        """Create overview sheet in Excel report"""
        # Title
        ws['A1'] = 'REPORTE DE ANÁLISIS DE ERRORES Y CALIDAD'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Generation info
        ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Metrics
        metrics = analysis['quality_metrics']
        
        row = 4
        ws[f'A{row}'] = 'MÉTRICAS PRINCIPALES'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        overview_data = [
            ['Total de documentos', analysis['total_documents']],
            ['Documentos con errores', analysis['documents_with_errors']],
            ['Tasa de éxito (%)', metrics.get('success_rate', 0)],
            ['Completitud promedio (%)', metrics.get('average_completeness', 0)],
            ['Puntuación de calidad', metrics.get('quality_score', 0)]
        ]
        
        for label, value in overview_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _create_errors_sheet(self, ws, analysis: Dict[str, Any]):
        """Create detailed errors sheet"""
        # Headers
        headers = ['Documento', 'Categoría', 'Severidad', 'Campo', 'Mensaje', 'Acción Sugerida', 'Impacto']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data
        row = 2
        for error in analysis['detailed_errors']:
            doc_index = error.get('document_index', 0) + 1
            
            ws.cell(row=row, column=1, value=f'Documento {doc_index}')
            ws.cell(row=row, column=2, value=self.error_categories[error['category']]['name'])
            ws.cell(row=row, column=3, value=self.severity_levels[error['severity']]['name'])
            ws.cell(row=row, column=4, value=error['field'])
            ws.cell(row=row, column=5, value=error['message'])
            ws.cell(row=row, column=6, value=error['suggested_action'])
            ws.cell(row=row, column=7, value=error['impact'])
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metrics_sheet(self, ws, analysis: Dict[str, Any]):
        """Create quality metrics sheet"""
        metrics = analysis['quality_metrics']
        
        # Title
        ws['A1'] = 'MÉTRICAS DE CALIDAD DETALLADAS'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Field detection rates
        ws[f'A{row}'] = 'TASAS DE DETECCIÓN POR CAMPO'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 2
        
        ws[f'A{row}'] = 'Campo'
        ws[f'B{row}'] = 'Tasa de Detección (%)'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        for field, rate in metrics.get('field_detection_rates', {}).items():
            field_name = self.validation_editor.field_definitions[field]['name']
            ws[f'A{row}'] = field_name
            ws[f'B{row}'] = round(rate, 1)
            row += 1
        
        row += 2
        
        # Severity distribution
        ws[f'A{row}'] = 'DISTRIBUCIÓN DE SEVERIDAD'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 2
        
        ws[f'A{row}'] = 'Severidad'
        ws[f'B{row}'] = 'Porcentaje (%)'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        for severity, percentage in metrics.get('severity_distribution', {}).items():
            ws[f'A{row}'] = self.severity_levels[severity]['name']
            ws[f'B{row}'] = round(percentage, 1)
            row += 1
    
    def _create_recommendations_sheet(self, ws, analysis: Dict[str, Any]):
        """Create recommendations sheet"""
        # Title
        ws['A1'] = 'RECOMENDACIONES PARA MEJORAR LA CALIDAD'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        for i, rec in enumerate(analysis['recommendations'], 1):
            ws[f'A{row}'] = f'{i}. {rec["title"]}'
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = f'Prioridad: {rec["priority"].upper()}'
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
            
            ws[f'A{row}'] = rec['description']
            row += 1
            
            ws[f'A{row}'] = 'Acciones recomendadas:'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for action in rec['actions']:
                ws[f'A{row}'] = f'• {action}'
                row += 1
            
            row += 1  # Empty row between recommendations