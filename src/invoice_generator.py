from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import io

class InvoiceGenerator:
    def __init__(self):
        # Static mock products data instead of loading from Excel
        self.products_db = pd.DataFrame([
            {"codigo": "SERV-001", "descripcion": "Servicio de Trámite Aduanal", "precio": 150.00},
            {"codigo": "LOG-001", "descripcion": "Logística y Transporte Local", "precio": 200.00},
            {"codigo": "ALM-001", "descripcion": "Almacenaje Fiscal", "precio": 75.50},
            {"codigo": "DOC-001", "descripcion": "Manejo de Documentación", "precio": 50.00},
            {"codigo": "IMP-001", "descripcion": "Impuestos de Importación", "precio": 0.00}, # Variable
        ])

    def load_products_from_template(self):
        """
        Returns the static DataFrame of mock products.
        """
        return self.products_db

    def generate_invoice_data(self, validated_data, selected_products):
        """
        Prepares the data dictionary for the invoice based on validated data and selected services.
        """
        invoice_items = []
        subtotal = 0.0
        
        # Add selected services/products
        for prod_name, info in selected_products.items():
            qty = info.get('quantity', 1)
            price = info.get('price', 0.0)
            total = qty * price
            subtotal += total
            
            # Find code from DB if possible
            code = "GEN-001"
            match = self.products_db[self.products_db['descripcion'] == prod_name]
            if not match.empty:
                code = match.iloc[0]['codigo']

            invoice_items.append({
                "codigo": code,
                "descripcion": prod_name,
                "cantidad": qty,
                "precio_unitario": price,
                "total": total
            })

        # Calculate taxes (15% ISV example)
        isv = subtotal * 0.15
        total_pagar = subtotal + isv

        # Prepare context
        client_name = validated_data.get("cliente_preview") or \
                      validated_data.get("bl_data", {}).get("consignee_details", {}).get("name", "CLIENTE GENÉRICO S.A.")
        
        rtn = validated_data.get("rtn_preview") or "00000000000000"
        
        address = validated_data.get("direccion_preview") or \
                  validated_data.get("bl_data", {}).get("consignee_details", {}).get("address", "CIUDAD GENÉRICA, PAÍS")

        context = {
            "cliente": client_name,
            "rtn": rtn,
            "direccion": address,
            "fecha": datetime.now().strftime("%d de %B de %Y"),
            "fecha_vencimiento": (datetime.now() + pd.DateOffset(months=1)).strftime("%d de %B de %Y"),
            "factura_n": "001-001-01-00000001",
            "cai": "000000-000000-000000-000000-000000-00",
            "bl_number": validated_data.get("bl_data", {}).get("bl_number", ""),
            "items": invoice_items,
            "subtotal": subtotal,
            "isv": isv,
            "total_pagar": total_pagar,
            "total_letras": self.number_to_text(total_pagar)
        }
        return context

    def create_invoice_excel(self, context):
        """
        Generates the Invoice Excel file programmmatically using openpyxl.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "FACTURA COMERCIAL"

        # Styles
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        bold_font = Font(name='Arial', size=10, bold=True)
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sección de Encabezado
        # Logo placeholder or Company Name
        ws['A1'] = "AGENCIA ADUANERA GENÉRICA"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F2')
        ws['A1'].alignment = center_align

        ws['A3'] = "R.T.N.: 08011900000000"
        ws['A3'].font = normal_font
        ws.merge_cells('A3:F3')
        ws['A3'].alignment = center_align

        ws['A4'] = "Dirección: Blvd. Suyapa, Tegucigalpa, Honduras"
        ws.merge_cells('A4:F4')
        ws['A4'].alignment = center_align

        # Cuadro de Información de Factura
        ws['E6'] = "FACTURA"
        ws['E6'].font = header_font
        ws['E6'].alignment = right_align
        ws['F6'] = context['factura_n']
        ws['F6'].font = normal_font
        ws['F6'].alignment = right_align

        ws['E7'] = "FECHA"
        ws['E7'].font = header_font
        ws['E7'].alignment = right_align
        ws['F7'] = context['fecha']
        ws['F7'].font = normal_font
        ws['F7'].alignment = right_align

        ws['E8'] = "CAI"
        ws['E8'].font = header_font
        ws['E8'].alignment = right_align
        ws['F8'] = context['cai']
        ws['F8'].font = Font(size=8)
        ws['F8'].alignment = right_align

        # Información del Cliente
        ws['A10'] = "CLIENTE:"
        ws['A10'].font = header_font
        ws['B10'] = context['cliente']
        ws['B10'].font = normal_font

        ws['A11'] = "R.T.N.:"
        ws['A11'].font = header_font
        ws['B11'] = context['rtn']
        ws['B11'].font = normal_font

        ws['A12'] = "DIRECCIÓN:"
        ws['A12'].font = header_font
        ws['B12'] = context['direccion']
        ws['B12'].font = normal_font

        ws['A13'] = "BL NO.:"
        ws['A13'].font = header_font
        ws['B13'] = context['bl_number']
        ws['B13'].font = normal_font

        # Encabezado de la Tabla de Ítems
        headers = ["CÓDIGO", "DESCRIPCIÓN", "CANTIDAD", "PRECIO UNITARIO", "TOTAL"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Datos de los Ítems
        current_row = 16
        for item in context['items']:
            ws.cell(row=current_row, column=1, value=item['codigo']).border = thin_border
            ws.cell(row=current_row, column=2, value=item['descripcion']).border = thin_border
            
            cell_qty = ws.cell(row=current_row, column=3, value=item['cantidad'])
            cell_qty.border = thin_border
            cell_qty.alignment = center_align

            cell_price = ws.cell(row=current_row, column=4, value=item['precio_unitario'])
            cell_price.border = thin_border
            cell_price.number_format = '#,##0.00'

            cell_total = ws.cell(row=current_row, column=5, value=item['total'])
            cell_total.border = thin_border
            cell_total.number_format = '#,##0.00'
            
            current_row += 1

        # Fill empty rows to make it look like a form (optional, simplified here)
        
        # Totales
        row_total = current_row + 2
        
        ws[f'D{row_total}'] = "SUBTOTAL"
        ws[f'D{row_total}'].font = bold_font
        ws[f'D{row_total}'].alignment = right_align
        ws[f'E{row_total}'] = context['subtotal']
        ws[f'E{row_total}'].number_format = '"L."#,##0.00'
        ws[f'E{row_total}'].font = bold_font

        row_total += 1
        ws[f'D{row_total}'] = "I.S.V. (15%)"
        ws[f'D{row_total}'].font = bold_font
        ws[f'D{row_total}'].alignment = right_align
        ws[f'E{row_total}'] = context['isv']
        ws[f'E{row_total}'].number_format = '"L."#,##0.00'
        ws[f'E{row_total}'].font = bold_font

        row_total += 1
        ws[f'D{row_total}'] = "TOTAL A PAGAR"
        ws[f'D{row_total}'].font = bold_font
        ws[f'D{row_total}'].alignment = right_align
        ws[f'E{row_total}'] = context['total_pagar']
        ws[f'E{row_total}'].number_format = '"L."#,##0.00'
        ws[f'E{row_total}'].font = bold_font
        ws[f'E{row_total}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Amount in words
        ws[f'A{row_total+2}'] = f"SON: {context['total_letras']}"
        ws[f'A{row_total+2}'].font = normal_font

        # Ancho de Columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10 # Extra column if needed

        # Output to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def number_to_text(self, number):
        """
        Simple number to text converter helper for Lempiras using generic logic.
        Simplified for prototype.
        """
        # A full implementation would be longer, using a placeholder for now
    def render_invoice_interface(self):
        import streamlit as st
        st.header("Generación de Factura Comercial")
        
        if 'validated_data' not in st.session_state or not st.session_state.validated_data:
            st.warning("No hay datos validados disponibles. Por favor, procese y valide documentos primero.")
            return

        validated_data = st.session_state.validated_data

        # Vista Previa de Información del Cliente
        with st.expander("Información del Cliente", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                st.info(f"**Cliente:** {validated_data.get('cliente_preview', 'No detectado')}")
                st.info(f"**RTN:** {validated_data.get('rtn_preview', 'No detectado')}")
            with col2:
                st.info(f"**Dirección:** {validated_data.get('direccion_preview', 'No detectado')}")
                st.info(f"**BL:** {validated_data.get('bl_data', {}).get('bl_number', 'No detectado')}")

        # Selección de Productos
        st.subheader("Servicios y Productos")
        
        products_df = self.load_products_from_template()
        product_names = products_df['descripcion'].tolist()
        
        # Initialize selection state if needed
        if 'invoice_selected_products' not in st.session_state:
            st.session_state.invoice_selected_products = {}

        # Product selector
        selected_product = st.selectbox("Agregar Servicio/Producto", [""] + product_names)
        
        if selected_product:
            col_qty, col_add = st.columns([1, 4])
            with col_qty:
                qty = st.number_input("Cantidad", min_value=1, value=1, key=f"qty_{selected_product}")
            with col_add:
                st.write("") # Spacer
                st.write("") # Spacer
                if st.button("Agregar", key=f"add_{selected_product}"):
                    row = products_df[products_df['descripcion'] == selected_product].iloc[0]
                    st.session_state.invoice_selected_products[selected_product] = {
                        'quantity': qty,
                        'price': float(row['precio'])
                    }
                    st.success(f"Agregado: {selected_product}")
                    st.rerun()

        # Display Selected Items
        if st.session_state.invoice_selected_products:
            st.markdown("### Ítems Seleccionados")
            
            selected_items = []
            total_preview = 0.0
            
            for name, details in st.session_state.invoice_selected_products.items():
                subtotal = details['quantity'] * details['price']
                total_preview += subtotal
                selected_items.append({
                    "Descripción": name,
                    "Cantidad": details['quantity'],
                    "Precio Unitario": f"L. {details['price']:.2f}",
                    "Total": f"L. {subtotal:.2f}"
                })
                
            st.table(selected_items)
            st.markdown(f"**Subtotal Estimado:** L. {total_preview:.2f}")

            if st.button("Limpiar Lista"):
                st.session_state.invoice_selected_products = {}
                st.rerun()

            # Botón Generar
            st.markdown("---")
            if st.button("Generar Factura", type="primary"):
                with st.spinner("Generando factura..."):
                    context = self.generate_invoice_data(validated_data, st.session_state.invoice_selected_products)
                    excel_bytes = self.create_invoice_excel(context)
                    
                    st.session_state.invoice_excel_bytes = excel_bytes
                    st.session_state.invoice_generation_success = True
                    st.rerun()
        else:
            st.info("Seleccione al menos un servicio para generar la factura.")

    def render_invoice_download(self):
        import streamlit as st
        st.success("Factura Generada Exitosamente")
        
        if 'invoice_excel_bytes' in st.session_state:
            st.download_button(
                label="Descargar Factura Excel",
                data=st.session_state.invoice_excel_bytes,
                file_name=f"Factura_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        if st.button("Generar Nueva Factura"):
            st.session_state.invoice_generation_success = False
            st.session_state.invoice_selected_products = {}
            if 'invoice_excel_bytes' in st.session_state:
                del st.session_state.invoice_excel_bytes
            st.rerun()
