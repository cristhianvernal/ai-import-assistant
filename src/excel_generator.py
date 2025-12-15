# Este es el excel_generator.py

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def create_final_excel_report(bl_data: Dict[str, Any], invoices_data: List[Dict[str, Any]], catalog_manager) -> io.BytesIO:
    """
    Genera el reporte final en Excel con 5 pestañas, replicando el formato
    solicitado por el cliente y usando fórmulas de Excel para cálculos clave.
    """
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active) # Eliminar la hoja por defecto

    # Estilos Reutilizables
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Pestaña 1: Aforo
    ws_aforo = wb.create_sheet(title="AFORO")

    # Encabezados
    aforo_headers = [
        "Item", "Bultos", "Contenido", "Marca", "Fob", "Flete", "Seg", "CIF",
        "Peso", "Aduana", "Posicion Arancelaria", "DAI", "SELECTIVO", "ISV"
    ]
    ws_aforo.append(aforo_headers)

    # Aplicar estilo a encabezados
    for cell in ws_aforo[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_thin

    # Consolidar todos los ítems de todas las facturas
    all_items = []
    for inv in invoices_data:
        all_items.extend(inv.get('items', []))

    # Llenar datos y fórmulas
    for i, item in enumerate(all_items, start=1):
        row_index = i + 1 # +1 porque la fila 1 es de encabezados
        
        # Fórmula para Seguro (1.5% del FOB)
        seguro_formula = f"=E{row_index}*0.015"
        
        # Fórmula para CIF (FOB + Flete + Seguro)
        cif_formula = f"=E{row_index}+F{row_index}+G{row_index}"
        
        arancel_code = catalog_manager.get_arancel_code(item.get('description_es', ''))

        row_data = [
            i,
            item.get('quantity'),
            item.get('description_es', item.get('description', 'No Desc')), # Contenido
            "S/M", # Marca
            item.get('fob_value'),
            item.get('freight_proportional'),
            seguro_formula,
            cif_formula,
            "", # Peso
            "", # Aduana
            arancel_code, # Posicion Arancelaria
            "20%", "0%", "18%" # Valores Fijos
        ]
        ws_aforo.append(row_data)

        # Aplicar formato a la fila
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws_aforo.cell(row=row_index, column=col_idx)
            cell.border = border_thin
            if col_idx in [5, 6, 7, 8]: # Columnas de moneda
                 cell.number_format = '"$"#,##0.00'
            if col_idx in [1, 2]: # Item y Bultos centrados
                 cell.alignment = Alignment(horizontal='center')


    # Ajustar ancho de columnas para AFORO
    column_widths_aforo = {'A': 5, 'B': 10, 'C': 50, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 10, 'J': 15, 'K': 20, 'L': 10, 'M': 12, 'N': 10}
    for col, width in column_widths_aforo.items():
        ws_aforo.column_dimensions[col].width = width


    # Pestaña 2: Listado DEVA
    ws_deva = wb.create_sheet(title="LISTADO DEVA")
    deva_headers = [
        "CODIGO DE BARRAS", "DESCRIPCION COMERCIAL", "MARCA", "MODELO/ESTILO",
        "PAIS DE ORIGEN", "UNIDADES", "PRECIO UNITARIO", "TOTAL"
    ]
    ws_deva.append(deva_headers)
    
    # Estilo encabezados
    for cell in ws_deva[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_thin

    # Llenar datos
    for item in all_items:
        total_price = item.get('quantity', 0) * item.get('unit_price', 0)
        row_data = [
            item.get('part_number', 'S/C'),
            item.get('description_es', ''),
            "S/M", "S/M", "CHINA",
            item.get('quantity'),
            item.get('unit_price'),
            total_price
        ]
        ws_deva.append(row_data)

    # Ajustar anchos
    column_widths_deva = {'A': 20, 'B': 50, 'C': 15, 'D': 15, 'E': 15, 'F': 10, 'G': 15, 'H': 15}
    for col, width in column_widths_deva.items():
        ws_deva.column_dimensions[col].width = width

    # Pestaña 3: Traducción
    ws_trad = wb.create_sheet(title="TRADUCCION")
    trad_headers = ["DESCRIPCION INGLES", "DESCRIPCION ESPAÑOL"]
    ws_trad.append(trad_headers)
    
    for cell in ws_trad[1]: # Estilo
        cell.font = header_font
        cell.fill = header_fill
    
    for item in all_items:
        ws_trad.append([item.get('description_original', ''), item.get('description_es', '')])

    column_widths_trad = {'A': 60, 'B': 60}
    for col, width in column_widths_trad.items():
        ws_trad.column_dimensions[col].width = width
        
    # Pestaña 4: Resumen Factura (Diseño Fijo)
    ws_resumen = wb.create_sheet(title="RESUMEN FACTURA")
    
    # Estilos específicos para esta hoja
    bold_font = Font(bold=True)
    
    # Contenido Fijo
    data_resumen = {
        "A1": "DATOS DEL EMBARQUE",
        "A3": "FACTURA No.",
        "A4": "FECHA",
        "A5": "PROVEEDOR",
        "A6": "BULTOS",
        "A7": "PESO",
        "A8": "VALOR FACTURA",
        "A9": "FLETE",
        "A10": "SEGURO",
        "A11": "VALOR CIF"
    }
    
    # Llenar datos fijos y aplicar negrita
    for cell_ref, value in data_resumen.items():
        ws_resumen[cell_ref] = value
        ws_resumen[cell_ref].font = bold_font

    # Llenar datos dinámicos
    first_invoice = invoices_data[0] if invoices_data else {}
    total_fob_shipment = sum(item.get('fob_value', 0) for item in all_items)
    total_freight_shipment = sum(item.get('freight_proportional', 0) for item in all_items)
    total_seguro_shipment = total_fob_shipment * 0.015
    total_cif_shipment = total_fob_shipment + total_freight_shipment + total_seguro_shipment
    
    ws_resumen['B3'] = first_invoice.get('invoice_number', 'N/A')
    ws_resumen['B4'] = first_invoice.get('invoice_date', 'N/A')
    ws_resumen['B5'] = bl_data.get('exporter_details', {}).get('name', 'N/A')
    ws_resumen['B6'] = bl_data.get('packages_count', 0)
    ws_resumen['B7'] = bl_data.get('gross_weight', 0)
    ws_resumen['B8'] = total_fob_shipment
    ws_resumen['B9'] = total_freight_shipment
    ws_resumen['B10'] = total_seguro_shipment
    ws_resumen['B11'] = total_cif_shipment
    
    # Aplicar formato de moneda
    for cell_ref in ['B8', 'B9', 'B10', 'B11']:
        ws_resumen[cell_ref].number_format = '"$"#,##0.00'
        
    ws_resumen.column_dimensions['A'].width = 20
    ws_resumen.column_dimensions['B'].width = 30

    # Pestaña 5: Catálogo
    ws_catalogo = wb.create_sheet(title="CATALOGO")
    catalogo_headers = ["Descripción Producto (Español)", "Posicion_Arancelaria"]
    ws_catalogo.append(catalogo_headers)

    for cell in ws_catalogo[1]: # Estilo
        cell.font = header_font
        cell.fill = header_fill

    for desc, code in catalog_manager.get_all_entries().items():
        ws_catalogo.append([desc, code])

    ws_catalogo.column_dimensions['A'].width = 60
    ws_catalogo.column_dimensions['B'].width = 30

    # Guardar el libro
    wb.save(output)
    output.seek(0)

    return output.getvalue()