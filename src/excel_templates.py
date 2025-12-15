# Este es el excel_templates.py

import pandas as pd
import io
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
import uuid
from openpyxl.utils import get_column_letter

# Simulación de un gestor de plantillas y sus cálculos
class ExcelTemplateManager:
    def __init__(self):
        self.templates = {
            'comercial': {
                'id': 'comercial',
                'name': 'Plantilla Comercial (Resumen)',
                'description': 'Reporte enfocado en valores FOB, CIF, y márgenes.',
                'columns': [
                    {'field': 'archivo', 'name': 'Documento Fuente', 'format': 'text'},
                    {'field': 'numero_bl', 'name': 'Nº BL', 'format': 'text'},
                    {'field': 'incoterm', 'name': 'INCOTERM', 'format': 'text'},
                    {'field': 'valor_factura', 'name': 'Valor Factura', 'format': 'currency'},
                    {'field': 'valor_fob', 'name': 'Valor FOB Consolidado', 'format': 'currency'},
                    {'field': 'valor_cif', 'name': 'Valor CIF Consolidado', 'format': 'currency'},
                    {'field': 'diff_fob_cif', 'name': 'Costo Total Flete/Seguro', 'format': 'currency', 'calc': 'CIF - FOB'},
                ]
            },
            'logistica': {
                'id': 'logistica',
                'name': 'Plantilla Logística (Peso/Volumen)',
                'description': 'Detalles de peso, bultos y costos por unidad.',
                'columns': [
                    {'field': 'numero_bl', 'name': 'Nº BL', 'format': 'text'},
                    {'field': 'numero_contenedor', 'name': 'Nº Contenedor', 'format': 'text'},
                    {'field': 'bultos', 'name': 'Bultos (Total)', 'format': 'integer'},
                    {'field': 'kilos', 'name': 'Kilos (Total)', 'format': 'decimal'},
                    {'field': 'valor_flete', 'name': 'Costo Flete BL', 'format': 'currency'},
                    {'field': 'cost_per_kg', 'name': 'Costo Flete por Kilo', 'format': 'currency', 'calc': 'Flete / Kilos'},
                ]
            }
        }
        
    def get_template(self, template_id: str) -> Dict[str, Any]:
        """Devuelve una plantilla por su ID."""
        return self.templates.get(template_id, self.templates['comercial'])

    def render_template_selector(self) -> str:
        """Renderiza la selección de plantilla en Streamlit."""
        options = {t['id']: f"{t['name']} ({t['description']})" for t in self.templates.values()}
        
        selected_key = st.selectbox(
            "Seleccione un tipo de plantilla:",
            options=list(options.keys()),
            format_func=lambda x: options[x],
            key="template_selector_box"
        )
        return selected_key

    def create_excel_with_template(self, processed_data: List[Dict[str, Any]], template_id: str) -> io.BytesIO:
        """
        Genera un archivo Excel basado en una plantilla predefinida.
        Recibe la data consolidada [BL/Factura]
        """
        output = io.BytesIO()
        template = self.get_template(template_id)
        
        wb = Workbook()
        
        # 1. Crear Pestaña de Datos (Detalle)
        self._create_template_data_sheet(wb, processed_data, template)
        
        # 2. Crear Pestaña de Resumen (Totales y Metadatos)
        # La función _create_template_summary es donde se origina el error MergedCell
        self._create_template_summary(wb, processed_data, template) 
        
        # Guardar en buffer
        wb.remove(wb.get_sheet_by_name('Sheet')) # Eliminar la hoja por defecto
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _safe_calc(self, data: Dict[str, Any], calc_rule: str) -> float:
        """Realiza cálculos seguros basados en reglas simples."""
        try:
            val_fob = data.get('valor_fob', 0.0) or 0.0
            val_cif = data.get('valor_cif', 0.0) or 0.0
            val_kilos = data.get('kilos', 0.0) or 0.0
            val_flete = data.get('valor_flete', 0.0) or 0.0

            if calc_rule == 'CIF - FOB':
                return float(val_cif) - float(val_fob)
            if calc_rule == 'Flete / Kilos':
                return float(val_flete) / float(val_kilos) if float(val_kilos) > 0 else 0.0
            
            return 0.0
        except (ValueError, TypeError, ZeroDivisionError):
            return 0.0
            
    def _create_template_data_sheet(self, wb, processed_data: List[Dict[str, Any]], template: Dict[str, Any]):
        """Crea la hoja principal de datos de la plantilla."""
        ws = wb.create_sheet(title=f"Datos_{template['id'].title()}")
        
        data_for_df = []
        
        for doc in processed_data:
            row_data = {}
            for col_def in template['columns']:
                field = col_def['field']
                
                if 'calc' in col_def:
                    row_data[col_def['name']] = self._safe_calc(doc, col_def['calc'])
                else:
                    row_data[col_def['name']] = doc.get(field, None)
            data_for_df.append(row_data)

        if not data_for_df:
            ws['A1'] = "No hay datos para esta plantilla."
            return

        df = pd.DataFrame(data_for_df)
        
        # Escribir el DataFrame en la hoja
        for r in dataframe_to_rows(df, header=True, index=False):
            ws.append(r)
        
        # Aplicar formato de encabezados y autoajuste básico
        self._apply_template_formatting(ws, template)
        
    def _apply_template_formatting(self, ws, template: Dict[str, Any]):
        """Aplica estilos básicos a la hoja de datos."""
        
        # Estilo de encabezado (Fila 1)
        header_row = ws[1]
        for cell in header_row:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Aplicar formato de moneda y decimales
        for col_index, col_def in enumerate(template['columns'], 1):
            col_letter = get_column_letter(col_index)
            
            if col_def['format'] == 'currency':
                format_code = '$#,##0.00'
            elif col_def['format'] == 'decimal':
                format_code = '0.00'
            else:
                format_code = None

            if format_code:
                for row_index in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row_index}'].number_format = format_code

        # Auto-ajuste de ancho de columna (simple)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width


    def _create_template_summary(self, wb, data: List[Dict[str, Any]], template: Dict[str, Any]):
        """Crea la hoja de resumen (totales, metadata, etc.)."""
        ws = wb.create_sheet(title="Resumen de Plantilla")
        
        # 1. Título y Metadata
        ws['A1'] = f'RESUMEN DE REPORTE: {template["name"].upper()}'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'ID de Plantilla: {template["id"]}'
        ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        
        ws.merge_cells('A1:D1')
        
        # 2. Resumen de Totales Clave
        row = 5
        ws[f'A{row}'] = 'MÉTRICAS CONSOLIDADAS'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Calcular totales
        total_kilos = sum(float(d.get('kilos', 0) or 0) for d in data)
        total_fob = sum(float(d.get('valor_fob', 0) or 0) for d in data)
        total_cif = sum(float(d.get('valor_cif', 0) or 0) for d in data)
        total_flete = sum(float(d.get('valor_flete', 0) or 0) for d in data)
        
        summary_data = [
            # CORRECCIÓN: Se agrega '0' como formato numérico para el conteo de documentos
            ['Total Documentos Procesados', len(data), '0'],
            ['Total Kilos', total_kilos, '0.00'],
            ['Total Valor FOB', total_fob, '$#,##0.00'],
            ['Total Valor CIF', total_cif, '$#,##0.00'],
            ['Total Costo Flete BL', total_flete, '$#,##0.00'],
            ['Costo Flete Promedio por Kilo', total_flete / total_kilos if total_kilos > 0 else 0, '$#,##0.0000']
        ]
        
        for label, value, fmt in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].number_format = fmt
            row += 1
            
        # Auto-adjust column widths (SOLUCIÓN DEL ERROR AQUÍ)
        for column in ws.columns:
            max_length = 0
            
            # Obtener la letra de la columna de forma segura
            column_letter = get_column_letter(column[0].column) # Usa el índice de columna para obtener la letra
            
            # Iterar sobre las celdas, ignorando las que forman parte de un rango fusionado
            for cell in column:
                try:
                    # Verifica si la celda NO es parte de un rango fusionado
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            break
                    
                    if not is_merged and cell.value is not None:
                         # Solo mide si la celda tiene contenido y no está fusionada
                         current_length = len(str(cell.value))
                         if current_length > max_length:
                             max_length = current_length
                             
                except AttributeError:
                    # Ignorar objetos que no son celdas estándar (MergedCell)
                    pass
            
            # Aplicar el ancho encontrado
            if max_length > 0:
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width